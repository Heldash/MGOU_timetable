# # -*- coding: cp1251 -*-
import os
import zipfile
import time
import openpyxl.cell.cell
from openpyxl import load_workbook
import json
import datetime
from pprint import pprint
import copy
# os.chdir(os.path.join("Fakultet","Медицинский факультет","Очная","Специалитет","31.05.01"))
# print(os.getcwd())
# print(os.getcwd())
def made_abrev(strok):
    arr_str = strok.split()
    arr_str = arr_str[:1]+[i[0] for i in arr_str[1:]]
    strok = "".join(arr_str)
    return strok

def tested_val(sheet,i,k,val):
    """Забирает значение смежной клетки"""
    for crange in sheet.merged_cells:
        clo, rlo, chi, rhi = crange.bounds
        top_value = sheet.cell(rlo,clo).value
        if rlo <= i and i <= rhi and clo <= (k) and (k) <= chi:
            val = top_value
            break
    return val
def make_gr_arr(arr:list):
    """Возвращает список с именами групп если есть копии то они именуется как название_группы*номер_группы"""
    res = []
    for i in range(len(arr)):
        if arr.count(arr[i])>=2:
            res.append(arr[i]+f"*{arr[:i].count(arr[i])+1}")
        else:
            res.append(arr[i])
    return res
def proverk_lesson(arr:list,pars:str,val:str):
    """Проверяет есть ли пара уже такая в расписании"""
    for i in range(len(arr)):
        if pars in arr[i] and val not in arr[i]:
            arr[i] = f"{arr[i]} | {val}"
            break
        elif pars in arr[i] and val in arr[i]:
            return sorted(arr)
    else:
        arr.append(f'{pars}: {val}')
    return sorted(arr)
def clear_str(strok):
    """Очищает строку от ненужных переносов и другий неприятностей форматирования"""
    #  
    arr_ahtung = [" ",",","\n","\t","|","<",">","/","\\","?",":",";","\"", "  "]

    for i in arr_ahtung:
        while i in strok:
            strok = strok.replace(i," ")
    return strok

def pars_lessons(sheet,dat_coord,pars_coord,special_timetable,col,groups,only_time = False):
    """Отсортировать список наооборот самый первый это самое большое число,потом по убывающей искать подходящее направление"""
    if groups == {}:
        date_pars = {}
    else:
        date_pars = groups
    act_dat = None
    if special_timetable:
        row = dat_coord[0]+2
    else:
        row = dat_coord[0]+1
    for j in range(row, sheet.max_row + 1):
        res = sheet.cell(row=j, column=dat_coord[1]).value
        if res == None:
            res = tested_val(sheet, j, col, res)
        if type(res) == datetime.datetime:
            res = res.strftime("%d.%m.%y")
            if res not in date_pars:
                date_pars[res] = []
            act_dat = res
            if only_time:
                cnt = 0
        if act_dat != None:
            if only_time:
                cnt+=1
            val = sheet.cell(row=j, column=col).value
            if val == None:
                val = tested_val(sheet, j, col, val)
            if val != None:
                para = sheet.cell(row=j, column=pars_coord[1]).value
                if para != None:
                    if type(para) == datetime.datetime:
                        para = para.strftime("%d.%m.%y")
                    if type(val) == datetime.datetime:
                        val = val.strftime("%d.%m.%y")
                    para = para.replace("\n", " ")
                    # para = para+": " +val
                    para = clear_str(para)
                    if only_time:
                        para = f"{cnt} пара {para}"
                    date_pars[act_dat] = proverk_lesson(date_pars[act_dat],para,val)
                    # date_pars[act_dat].append(para)
    for dat in copy.copy(date_pars):
        if date_pars[dat] == []:
            del date_pars[dat]
    return date_pars

def upload_rasp():
    osn_dir = os.path.abspath(os.getcwd())
    # try:
    with open("mgou.json","r") as S:
        osnova = json.load(S)
    for fak in osnova:
        for form_obych in osnova[fak]:
            for level in osnova[fak][form_obych]:
                names_napr = list(osnova[fak][form_obych][level].keys())
                names_napr.sort(key=len,reverse=True)
                name_abreviature = {}
                all_groups = {}
                for temp in names_napr:
                    all_groups[" ".join(temp.split())] = {"ФТД":{"Занятия": {},
                              "Экзамены":{},
                              "Зачеты":{},
                              "Гос.экз":{}},
                              "Учебные дисциплины": {"Занятия": {},
                                      "Экзамены": {},
                                      "Зачеты": {},
                                      "Гос.экз": {}}
                              }
                    name_abreviature[made_abrev(clear_str(temp))] = temp # В дальнейшем нужно для сопоставления расписания с направлением
                # print(name_abreviature)
                is_eror = False
                # print(list(all_groups.keys()))
                put_dir = os.path.join(osn_dir,"Fakultet",fak,form_obych,level)
                files = [i for i in os.listdir(put_dir) if i[-4:]=="xlsx" and i[0]!="~"]
                if files == []:
                    continue
                title = None
                for i in files:
                    try:
                        wb = load_workbook(f"{put_dir}/{i}")
                    # except zipfile.BadZipfile as e:
                    #     try:
                    #         for url in osnova[fak][form_obych][level][prog]:
                    #             download_file(url.split("/")[5], i[-5])
                    #         wb = load_workbook(f"{put_dir}/{i}")
                    #     except:
                    #         continue
                    except:
                        continue
                    sheet = wb[wb.sheetnames[0]]
                    title = sheet.cell(row = 1,column = 1).value
                    ser_row,ser_col = 1,1 #Искомые строки и столбцы от search row and search column
                    while type(title) != str or type(title) == datetime.datetime or "Расписание" not in title:
                        if ser_col>sheet.max_column:
                            ser_row +=1
                            ser_col =1
                        title = sheet.cell(row = ser_row, column = ser_col).value
                        if title == None:
                            title = tested_val(sheet, ser_row, ser_col, title)
                        # if type(title) == str or type(title) != datetime.datetime or "Расписание" in title:
                        #     continue
                        ser_col+=1
                        if ser_row >10:
                            is_eror = True
                            break
                    if title == None or type(title) == datetime.datetime:
                        continue
                    # if title ==None:
                    #     title = tested_val(sheet,1,3,title)
                    # if ser_row ==
                    actual_group = name_abreviature[(i.split())[0]]
                    title = clear_str(title)

                    group_row = ser_row +1
                    day_coord = (0, 0)
                    pars_coord = (0, 0)
                    ro = co = 1
                    only_time = False
                    while day_coord == (0, 0) and pars_coord == (0, 0):
                        if type(sheet.cell(ro,co))==openpyxl.cell.cell.MergedCell:
                            val = tested_val(sheet,ro,co,(sheet.cell(ro,co).value)).lower()
                        else:
                            try:
                                val = (sheet.cell(ro,co).value).lower()
                            except:
                                val = sheet.cell(ro,co).value
                        if val is not None:
                            if any(i in val for i in ("дата","день","день недели")):
                                day_coord = (ro,co)
                            # elif "пара"in val:
                            #     pars_coord = (ro,co)
                            #     only_time = False
                            # elif "время" in val:
                            #     pars_coord = (ro, co)
                            #     only_time = True

                            elif type(val) == datetime.datetime:
                                group_row = day_coord[0]
                                if type(sheet.cell(day_coord[0], day_coord[1]+1)) == openpyxl.cell.cell.MergedCell:
                                    val_par = str(tested_val(sheet, day_coord[0], day_coord[1]+1, (sheet.cell(day_coord[0], day_coord[1]+1).value))).lower()
                                else:
                                    try:
                                        val_par = str(sheet.cell(day_coord[0], day_coord[1]+1).value).lower()
                                    except:
                                        val_par = sheet.cell(day_coord[0], day_coord[1]+1).value
                                if "пара" in val_par:
                                    only_time = False
                                elif "время" in val_par:
                                    only_time = True
                                pars_coord = (day_coord[0],day_coord[1]+1)
                        if ro > 10:
                            break
                        ro +=1
                    if day_coord == (0, 0) or pars_coord == (0, 0):
                        continue


                    special_timetable = False #Необычное ли расписание
                    if actual_group != None:
                        col = 1
                        day_coord = (group_row,1)
                        pars_coord = (group_row,2)

                        if type(sheet.cell(row = group_row+1,column = 1).value) == datetime.datetime or \
                                type(sheet.cell(row = group_row+1,column = 1).value) == str or \
                                type(sheet.cell(row = group_row+1,column = 2).value) == str:
                            special_timetable = False
                        elif sheet.cell(row = group_row+1,column = 1).value== None:
                            special_timetable = True
                        while sheet.cell(row = group_row,column = col).value != None \
                                or tested_val(sheet,group_row,col,sheet.cell(row = group_row,column = col).value):
                            if str(sheet.cell(row = group_row,column = col).value).lower() == "день" \
                                    or "день" in str(sheet.cell(row = group_row,column = col).value).lower():
                                day_coord = (group_row,col)
                                col+=1
                                continue
                            elif str(sheet.cell(row = group_row,column = col).value).lower() == "пара" \
                                    or "пара" in str(sheet.cell(row = group_row,column = col).value).lower():
                                pars_coord = (group_row, col)
                                col += 1
                                continue
                            elif all(i not in str(sheet.cell(row = group_row,column = col).value).lower()
                                     for i in ("день","пара")):

                                if special_timetable:
                                    group_name = sheet.cell(row=group_row+1, column=col).value
                                    if group_name == None:
                                        group_name = tested_val(sheet,group_row+1,col,group_name)
                                        if group_name == None:
                                            group_name = sheet.cell(row = group_row,column = col).value
                                else:
                                    group_name = sheet.cell(row = group_row,column = col).value
                                    if group_name == None:
                                        group_name = tested_val(sheet,group_row,col,group_name)
                                if group_name != None and type(group_name)!=datetime.datetime:
                                    group_name = clear_str(group_name)
                                    group_name = group_name.replace(" ","")
                                    # Далее классифицируется таблица по типу и если уже была группа в списке то передаем список с группой если нет то пустой
                                    if "занятий" in title:
                                        if "факультативным" in title:
                                            if group_name not in all_groups[actual_group]["ФТД"]["Занятия"]:
                                                all_groups[actual_group]["ФТД"]["Занятия"][group_name] = pars_lessons(
                                                    sheet,day_coord,pars_coord,special_timetable,col,{})
                                            else:
                                                all_groups[actual_group]["ФТД"]["Занятия"][group_name] = pars_lessons(sheet,day_coord,
                                                                                                                      pars_coord,special_timetable,
                                                                                                                      col,all_groups[actual_group]["ФТД"]["Занятия"][group_name],
                                                                                                                      only_time)

                                        else:
                                            if group_name not in all_groups[actual_group]["Учебные дисциплины"]["Занятия"]:
                                                all_groups[actual_group]["Учебные дисциплины"]["Занятия"][group_name] = pars_lessons(sheet,day_coord,pars_coord,special_timetable,col,{},
                                                                                                                                     only_time)
                                            else:
                                                all_groups[actual_group]["Учебные дисциплины"]["Занятия"][
                                                    group_name] = pars_lessons(sheet, day_coord, pars_coord,
                                                                    special_timetable, col,
                                                                    all_groups[actual_group]["Учебные дисциплины"]["Занятия"][group_name],only_time)
                                    elif "экзаменационной" in title:
                                        if "факультативным" in title:
                                            if group_name not in all_groups[actual_group]["ФТД"]["Экзамены"]:
                                                all_groups[actual_group]["ФТД"]["Экзамены"][group_name] = pars_lessons(sheet,day_coord,pars_coord,special_timetable,col,{},only_time)
                                            else:
                                                all_groups[actual_group]["ФТД"]["Экзамены"][group_name] = pars_lessons(
                                                    sheet,day_coord,pars_coord,special_timetable,col,all_groups[actual_group]["ФТД"]["Экзамены"][group_name],only_time)

                                        else:
                                            if group_name not in all_groups[actual_group]["Учебные дисциплины"]["Экзамены"]:
                                                all_groups[actual_group]["Учебные дисциплины"]["Экзамены"][group_name] = pars_lessons(sheet,
                                                                                                                                      day_coord,pars_coord,special_timetable,col,
                                                                                                                                      {},only_time)
                                            else:
                                                all_groups[actual_group]["Учебные дисциплины"]["Экзамены"][group_name] = pars_lessons(sheet,
                                                                                                                        day_coord,pars_coord,special_timetable,col,
                                                                                                                        all_groups[actual_group]["Учебные дисциплины"]["Экзамены"][group_name],
                                                                                                                                      only_time)
                                    elif"зачетной" in title:
                                        if "факультативным" in title:
                                            if group_name not in all_groups[actual_group]["ФТД"]["Зачеты"]:
                                                all_groups[actual_group]["ФТД"]["Зачеты"][group_name] = pars_lessons(sheet,day_coord,
                                                                                                                     pars_coord,special_timetable,col,{},only_time)
                                            else:
                                                all_groups[actual_group]["ФТД"]["Зачеты"][group_name] = pars_lessons(
                                                    sheet, day_coord,
                                                    pars_coord, special_timetable, col, all_groups[actual_group]["ФТД"]["Зачеты"][group_name],only_time)

                                        else:
                                            if group_name not in all_groups[actual_group]["Учебные дисциплины"]["Зачеты"]:
                                                all_groups[actual_group]["Учебные дисциплины"]["Зачеты"][group_name] = pars_lessons(sheet,day_coord,
                                                                                                                     pars_coord,special_timetable,col,{},only_time)
                                            else:
                                                all_groups[actual_group]["Учебные дисциплины"]["Зачеты"][group_name] = pars_lessons(sheet, day_coord,
                                                                               pars_coord, special_timetable, col, all_groups[actual_group]["Учебные дисциплины"]["Зачеты"][group_name],
                                                                                                                                    only_time)
                                    elif"аттестационных испытаний" in title:
                                        if "факультативным" in title:
                                            if group_name not in all_groups[actual_group]["ФТД"]["Гос.экз"]:
                                                all_groups[actual_group]["ФТД"]["Гос.экз"][group_name] = pars_lessons(sheet,day_coord,pars_coord,special_timetable,col,{},only_time)
                                            else:
                                                all_groups[actual_group]["ФТД"]["Гос.экз"][group_name] = pars_lessons(
                                                    sheet,day_coord,pars_coord,special_timetable,col,all_groups[actual_group]["ФТД"]["Гос.экз"][group_name],only_time)

                                        else:
                                            if group_name not in all_groups[actual_group]["Учебные дисциплины"]["Гос.экз"]:
                                                all_groups[actual_group]["Учебные дисциплины"]["Гос.экз"][group_name] = pars_lessons(sheet,day_coord,pars_coord,special_timetable,col,{},only_time)
                                            else:
                                                all_groups[actual_group]["Учебные дисциплины"]["Гос.экз"][
                                                    group_name] = pars_lessons(sheet, day_coord, pars_coord,
                                                                               special_timetable, col, all_groups[actual_group]["Учебные дисциплины"]["Гос.экз"][group_name],only_time)
                            col+=1
                if os.path.exists(os.path.join(put_dir,"group.json")):
                    with open(f"{put_dir}/group.json","r") as file:
                        try:
                            timetabl = json.load(file)
                        except:
                            timetabl = {}
                    if timetabl == {}:
                        with open(f"{put_dir}/group.json", "w") as f:
                            json.dump(all_groups, f, indent=2, ensure_ascii=True)
                            print(f"Создан новый файл json")
                    else:
                        with open(f"{put_dir}/group.json", "w") as f:
                            for i in all_groups:
                                for k in all_groups[i]["ФТД"]:
                                    if all_groups[i]["ФТД"][k] != {}:
                                        timetabl[i]["ФТД"][k] = all_groups[i]["ФТД"][k]
                                for j in all_groups[i]["Учебные дисциплины"]:#Доделать проверку
                                    if all_groups[i]["Учебные дисциплины"][j] != {}:
                                        timetabl[i]["Учебные дисциплины"][j] = all_groups[i]["Учебные дисциплины"][j]
                            json.dump(timetabl,f,ensure_ascii=True,indent=2)
                            print(f"Обновлен json")

                else:
                    with open(f"{put_dir}/group.json","w") as f:
                        json.dump(all_groups,f,indent=2, ensure_ascii=True)
                        print(f"Создан новый файл json")
if __name__ == "__main__":
    # wb = load_workbook("./Fakultet/Медицинский факультет/Очная/Специалитет/31.05.01Лд 1.xlsx")
    # sheet = wb[wb.sheetnames[0]]
    # print(sheet.cell(4,1))
    # print(tested_val(sheet,4,1,sheet.cell(4,1)))
    start = time.time()
    upload_rasp()
    fin = time.time()

    print(fin-start)
