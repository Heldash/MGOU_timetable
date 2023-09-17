# # -*- coding: cp1251 -*-

import os
import json
import re

import openpyxl.cell
from openpyxl import load_workbook
strok = "01 ПООМ'21 СНвПИиО-Д.1"
# res = re.fullmatch(r"(\d*)([^0-9A-Я])(\D*)([^0-9A-Я])(\d*)([^0-9A-Я])(\D*)([^0-9A-Я])(\d*)",strok)
# # res = re.sub()
# print(res)
# res = re.sub(r"(\d)*[^0-9A-Я/](\D*)[^0-9A-Я/](\d*)[^0-9A-Я/](\D*)[^0-9A-Я/](\d*)[^0-9A-Я/](\d*)",r"\1.\2.\3.\4.\5*\6",strok)
# print(res)
# strok = "занят"
# print(re.findall(r"\bзанят.?.?.?\b",strok))
# # if res:
# #     print("sss")
# # a = [(1,2),(3,4)]
# # print(isinstance(a,list))
# # wb = load_workbook("Fakultet/Юридический факультет/Очная/Бакалавр/40.03.01ЮПдсииия 28.xlsx")
# # sheet = wb[wb.sheetnames[0]]
# # val = sheet.merged_cells
# # yai = sheet.cell(1,3)
# # if isinstance(yai,openpyxl.cell.Cell):
# #     print('sssssssssssssssssssssssssssssssss')
#
# # a = ["Rib","ali","wam"]
# # for i,j in enumerate(a):
# #     print(i,"$",j)
# # with open("test.json","r") as f:
# #     a = json.load(f)
# #     print(len(a.keys()))
#
# def read_merge_cell(val,sh_merge_cells):
#     for i in sh_merge_cells:
#         if val.coordinate in i:
#             return i.start_cell
# def is_merg(val):
#     if type(val) == openpyxl.cell.MergedCell:
#         return True
#     else:
#         return False
#
# def read_cell(sheet,row,col,merged_cells):
#     """Читает любую яйчейку и возвращает ее значание"""
#     val = sheet.cell(row,col)
#     if is_merg(val):
#         return read_merge_cell(val,merged_cells).value
#     else:
#         return val.value
# wb = load_workbook("Fakultet/Медицинский факультет/Очная/Специалитет/31.05.01Лд 2.xlsx")
# sheet = wb[wb.sheetnames[0]]
# all_merg = sheet.merged_cells
# print(type(read_cell(sheet,5,1,all_merg)))
# month = {"января":1,
#          "февраля":2,
#          "марта":3,
#          "апреля":4,
#          "мая":5,
#          "июня":6,
#          "июля":7,
#          "августа":8,
#          "сентября":9,
#          "октября":10,
#          "ноября":11,
#          "декабря":12}
# w_day = {"понедельник":1,
#          "вторник":2,
#          "среда":3,
#          "четверг":4,
#          "пятница":5,
#          "суббота":6,
#          "воскресенье":7}
# day = input()
# day = day.split()
# for i in day:
#     if i in w_day:
#         week_day = w_day[i]
#     elif i in month:
#         mesyac = month[i]
#     elif i.isdigit():
#         d = int(i)
# print(f"{d}.{mesyac}.{2023}")
with open("proffesors.json", "r") as file:
    proffessors = json.load(file)
def clear_fio(pairs: str):
    finded = re.findall(r"[А-Яа-яЁё]+.?[А-Я]\.[А-Я]\.", pairs)
    temp = re.findall(r"[А-Яа-яЁё]+.?[А-Я]\.\b", pairs)

    # if temp:
    #     finded += temp
    print(finded)
    if finded:
        if len(finded) == 1:
            if finded[0] in proffessors:
                pairs = pairs.replace(finded[0], "")
        if len(finded) > 1:
            for i in range(len(finded)):
                if finded[i] in proffessors:
                    if i == len(finded) - 1:
                        pairs = pairs.replace(f"{finded[i]}", "")
                    else:
                        pairs = pairs.replace(f"{finded[i]},", "")
        while "  " in pairs:
            pairs = pairs.replace("  ", " ")
    else:
        finded = re.findall(r"[А-Яа-яЁё]+.[А-Я]\.[А-Я]\.", pairs)
        if finded:
            print(f"найдено необычное{finded}")
    if any(i in pairs for i in temp) and any(k in proffessors for k in temp):
        if len(temp) == 1:
            if temp[0] in proffessors:
                pairs = pairs.replace(temp[0], "")
        if len(temp) > 1:
            for i in range(len(temp)):
                if temp[i] in proffessors:
                    if i == len(finded) - 1:
                        pairs = pairs.replace(f"{temp[i]}", "")
                    else:
                        pairs = pairs.replace(f"{temp[i]},", "")
        while "  " in pairs:
            pairs = pairs.replace("  ", " ")
    return pairs
# print(re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\D+[^0-9A-Я]\d+", strok))
# clear_fio("Мусатова М.А.")
def f(x):
    return x*2
print(f(4))
