# # -*- coding: cp1251 -*-
import datetime
import json
import os
import openpyxl.cell
from openpyxl import load_workbook
import re
import time

with open("proffesors.json", "r") as file:
    proffessors = json.load(file)

def read_merge_cell(val, sh_merge_cells):
    for i in sh_merge_cells:
        if val.coordinate in i:
            return i.start_cell


def is_merg(val):
    if type(val) == openpyxl.cell.MergedCell:
        return True
    else:
        return False


def read_cell(sheet, row, col, merged_cells):
    """Читает любую яйчейку и возвращает ее значание"""
    val = sheet.cell(row, col)
    if is_merg(val):
        return read_merge_cell(val, merged_cells).value
    else:
        return val.value


def made_abrev(strok):
    """Создает абревиатуру из первых букв каждого слова"""
    arr_str = strok.split()
    arr_str = arr_str[:1] + [i[0] for i in arr_str[1:]]
    strok = "".join(arr_str)
    return strok


def convert_to_date(strok: str):
    month = {"января": 1,
             "февраля": 2,
             "марта": 3,
             "апреля": 4,
             "мая": 5,
             "июня": 6,
             "июля": 7,
             "августа": 8,
             "сентября": 9,
             "октября": 10,
             "ноября": 11,
             "декабря": 12}
    w_day = {"понедельник": 1,
             "вторник": 2,
             "среда": 3,
             "четверг": 4,
             "пятница": 5,
             "суббота": 6,
             "воскресенье": 7}
    res = None
    day = None
    mes = None
    for i in strok.split():
        if i in month:
            mes = month[i]
        elif i.isdigit():
            day = int(i)
    if day is not None and mes is not None:
        res = datetime.date(datetime.datetime.now().timetuple().tm_year, mes, day)
    return res


def clear_str(strok):
    """Очищает строку от ненужных переносов и другий неприятностей форматирования"""
    #  
    arr_ahtung = [" ", ",", "\n", "\t", "|", "<", ">", "/", "\\", "?", ":", ";", "\"", "  "]

    for i in arr_ahtung:
        while i in strok:
            strok = strok.replace(i, " ")
    return strok


def clear_fio(pairs: str):
    finded = re.findall(r"[А-Яа-яЁё]+.?[А-Я]\..?[А-Я]\.", pairs)

    find_sok = re.findall(r"[А-Яа-яЁё]+.?[А-Я]\..?\b", pairs)
    sokr = {}
    # if temp:
    #     finded += temp
    # print(finded)
    temp = {}
    if finded:
        for i in finded:
            temp["".join(i.split())] = i
        if len(finded) == 1:
            if temp[list(temp.keys())[0]] in proffessors:
                pairs = pairs.replace(temp[list(temp.keys())[0]], "")
        if len(finded) > 1:
            for ind,name in enumerate(temp):
                if name in proffessors:
                    if ind == len(finded) - 1:
                        pairs = pairs.replace(f"{temp[name]}", "")
                    else:
                        pairs = pairs.replace(f"{temp[name]},", "")
        while "  " in pairs:
            pairs = pairs.replace("  ", " ")
    else:
        finded = re.findall(r"[А-Яа-яЁё]+.[А-Я]\.[А-Я]\.", pairs)
        if finded:
            print(f"найдено необычное{finded}")

    if sokr:
        for i in find_sok:
            sokr["".join(i.split())] = i
        if any(sokr[i] in pairs for i in sokr) and any("".join(k.split) in proffessors for k in sokr):
            if len(sokr) == 1:
                if sokr[list(sokr.keys())[0]] in proffessors:
                    pairs = pairs.replace(sokr[list(sokr.keys())[0]], "")
            if len(sokr) > 1:
                for ind,name in enumerate(sokr):
                    if sokr[name] in proffessors:
                        if ind == len(finded) - 1:
                            pairs = pairs.replace(f"{name}", "")
                        else:
                            pairs = pairs.replace(f"{name},", "")
            while "  " in pairs:
                pairs = pairs.replace("  ", " ")
    return pairs


def pars_line(sheet, row: int, day_col, pairs_col, gr: list, all_merg) -> dict:
    """Считывает пары по линиям"""
    groups = {i: {} for i in sorted(set(gr))}
    for i in range(row, sheet.max_row + 1):
        temp_group = {}
        day = read_cell(sheet, i, day_col, all_merg)
        if isinstance(day, str):
            res = convert_to_date(day)
            if res is None:
                continue
            elif res is not None:
                day = res.strftime("%d.%m.%y")
        if isinstance(day, datetime.datetime):
            day = day.strftime("%d.%m.%y")
        pairs = read_cell(sheet, i, pairs_col, all_merg)
        if pairs is not None:
            pairs = " ".join(pairs.split())
        start_col = pairs_col + 1
        if day is None and pairs is None:
            continue
        for ind, name in enumerate(gr):
            val = read_cell(sheet, i, start_col + ind, all_merg)
            if isinstance(val, datetime.datetime):
                start_col += 2
                val = read_cell(sheet, i, start_col + ind, all_merg)
            if "пара" in str(val).lower():
                start_col += 1
                val = read_cell(sheet, i, start_col + ind, all_merg)
            if val is not None and str(val).split() != []:
                if len(str(val).split())> 1:
                    val = " ".join(val.split())
                else:
                    while "\n" in val:
                        val = val.replace("\n", " ")
                    while "  " in val:
                        val = val.replace("  ", " ")
                val = clear_fio(val)
                if name not in temp_group:
                    temp_group[name] = " ".join(val.split())
                elif val not in temp_group[name].split(" | "):
                    temp_group[name] = temp_group[name] + " | " + " ".join(val.split())
        for name_gr in temp_group:
            if day not in groups[name_gr]:
                groups[name_gr][day] = []
            groups[name_gr][day].append(f"{pairs} : {temp_group[name_gr]}")
    return groups


def upload_rasp():
    with open("mgou.json", "r") as f:
        mgou = json.load(f)
    with open("all_timetable.json", "r") as file:
        try:
            all_timetable = json.load(file)
        except:
            all_timetable = {fak: {form: {lev: {clear_str(naprav): {"ФТД": {"Занятия": {},
                                                                            "Экзамены": {},
                                                                            "Зачеты": {},
                                                                            "Гос.экз": {}
                                                                            },
                                                                    "Учебные дисциплины": {"Занятия": {},
                                                                                           "Экзамены": {},
                                                                                           "Зачеты": {},
                                                                                           "Гос.экз": {}
                                                                                           }} for naprav in
                                                mgou[fak][form][lev]}
                                          for lev in mgou[fak][form]} for form in mgou[fak]} for fak in mgou}
    main_dir = os.path.join(os.getcwd(), "Fakultet")
    month = {"января": 1,
             "февраля": 2,
             "марта": 3,
             "апреля": 4,
             "мая": 5,
             "июня": 6,
             "июля": 7,
             "августа": 8,
             "сентября": 9,
             "октября": 10,
             "ноября": 11,
             "декабря": 12}
    w_day = {"понедельник": 1,
             "вторник": 2,
             "среда": 3,
             "четверг": 4,
             "пятница": 5,
             "суббота": 6,
             "воскресенье": 7}
    for fak in mgou:
        print(fak)
        for form in mgou[fak]:
            for level in mgou[fak][form]:
                work_dir = os.path.join(main_dir, fak, form, level)
                arr_exl = [i for i in os.listdir(work_dir) if i[-4:] == "xlsx" and i[0] != "~"]
                naprav = {clear_str(k): [] for k in mgou[fak][form][level]}
                keyword = {made_abrev(clear_str(k)): clear_str(k) for k in mgou[fak][form][level]}
                for i in arr_exl:
                    name = i.split()[0]
                    try:
                        if keyword[name] in naprav:
                            naprav[keyword[name]].append(i)
                    except:
                        print("-_-_-_-_-_-_-_-_EROR_-_-_-_-_-_-_-_-_-_-_-_-_-_-")
                        print(work_dir)
                        print(naprav)
                        print(arr_exl)
                        print(keyword[name])
                        exit()

                if arr_exl == []:
                    continue
                for nap in naprav:
                    pred_title = None
                    # if nap not in all_timetable[fak][form][level]:
                    for exel in naprav[nap]:
                        try:
                            wb = load_workbook(os.path.join(work_dir, exel))
                        except:
                            continue
                        sheet = wb[wb.sheetnames[0]]
                        all_merg = sheet.merged_cells
                        info_table = []  # [0] - учебные или факультативные, [1] - занятия,экзамен,зачет
                        title = None
                        for row in range(1, sheet.max_row):
                            for col in range(1, sheet.max_column):
                                val = read_cell(sheet, row, col, all_merg)
                                if val and type(val) == str:
                                    if "расписание" in val.lower():
                                        title = clear_str(val.lower())
                                        pred_title = clear_str(val.lower())
                                        break
                            if title is not None:
                                break
                            if row > 10:
                                break
                        if title is None and pred_title is None:
                            continue
                        elif title is None and pred_title is not None:
                            title = pred_title

                        if "факультативным" in title:
                            info_table.append("ФТД")
                        else:
                            info_table.append("Учебные дисциплины")

                        if re.findall(r"\bзанят.?.?.?\b", title):  # "занятий" in title:
                            info_table.append("Занятия")
                        elif re.findall(r"\bэкзаменационн.?.?.?\b", title):
                            info_table.append("Экзамены")
                        elif re.findall(r"\bзачет.?.?.?.?\b", title):
                            info_table.append("Зачеты")
                        elif "аттестационных испытаний" in title:
                            info_table.append("Гос.экз")

                        days_coord = pairs_coord = (0, 0)
                        for row in range(1, sheet.max_row):
                            for col in range(1, sheet.max_column):
                                val = read_cell(sheet, row, col, all_merg)
                                if val and type(val) == str:
                                    val = val.lower()
                                    if "день" in val or "дата" in val:
                                        days_coord = (row, col)
                                    elif "время" in val or "пара" in val:
                                        pairs_coord = (row, col)
                                elif isinstance(val,datetime.datetime):
                                    days_coord = (row-1,col)
                                    pairs_coord = (row-1,col+1)
                                elif convert_to_date(str(val)) is not None:
                                    days_coord = (row - 1, col)
                                    pairs_coord = (row - 1, col + 1)
                                if days_coord != (0, 0) and pairs_coord != (0, 0):
                                    break
                                if row > 7:
                                    break
                            if days_coord != (0, 0) and pairs_coord != (0, 0):
                                break
                            if row > 7:
                                break
                        if days_coord == (0, 0) and pairs_coord == (0, 0):
                            continue
                        elif pairs_coord != (0,0) and days_coord == (0,0):
                            days_coord = (pairs_coord[0],pairs_coord[1]-1)
                        elif days_coord != (0, 0) and pairs_coord == (0, 0):
                            pairs_coord = (days_coord[0], days_coord[1] + 1)
                        group_row = days_coord[0]
                        groups = []
                        # print(days_coord,pairs_coord)
                        # if read_cell(sheet,pairs_coord[0],pairs_coord[1]+1,all_merg) is not None:
                        for col in range(pairs_coord[1] + 1, sheet.max_column + 1):
                            # try:
                            val = read_cell(sheet, group_row, col, all_merg)
                            if val is not None and type(val) != datetime.datetime and type(val) == str:
                                if re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\D+[^0-9A-Я]\d+",
                                              str(val)) or re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\d+",
                                                                      str(val)):
                                    if " " in val:
                                        while " " in val:
                                            val = val.replace(" ", "")
                                    # if "разговорный" in val.lower():
                                    val = val.split("\n")[0]
                                    groups.append(val)
                                else:
                                    print(f"не подошло{val}")
                            # except:
                            # print(work_dir,exel)
                        while groups == []:
                            group_row += 1
                            # if read_cell()
                            if read_cell(sheet, group_row, days_coord[1],
                                         all_merg) != datetime.datetime or convert_to_date(
                                    str(read_cell(sheet, group_row, days_coord[1], all_merg))) is not None:
                                for col in range(pairs_coord[1] + 1, sheet.max_column + 1):
                                    # try:
                                    val = read_cell(sheet, group_row, col, all_merg)
                                    if val is not None and type(val) != datetime.datetime and type(val) == str:
                                        if re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\D+[^0-9A-Я]\d+",
                                                      str(val)) or re.findall(
                                            r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\d+", str(val)) or re.findall(
                                            r"\d+[^0-9A-Я]\D+[^0-9A-Я]+\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\d+",
                                            str(val)):
                                            # if " " in val:
                                            #     while " " in val:
                                            #         val = val.replace(" ", "")
                                            val = val.split("\n")[0]
                                            groups.append(val)
                                        else:
                                            print(f"не подошло{val}")
                            else:
                                break
                        if groups is []:
                            print(days_coord, pairs_coord)
                            print(work_dir, exel)
                            # print(re.findall(r"\d*\.\D*\.\d*\.\D*\.\d*", str(val)))
                            print(read_cell(sheet, pairs_coord[0], pairs_coord[1] + 1, all_merg))
                            print(type(read_cell(sheet, pairs_coord[0], pairs_coord[1] + 1, all_merg)))
                            continue

                        if isinstance(read_cell(sheet, group_row + 1, days_coord[1], all_merg), datetime.datetime) or \
                                "пара" in str(read_cell(sheet, group_row + 1, pairs_coord[1], all_merg)).lower() \
            or convert_to_date(str(read_cell(sheet, group_row + 1, days_coord[1], all_merg))) is not None:
                            try:

                                if "Медицинский" in work_dir:
                                    print("simp")
                                pair = pars_line(sheet, group_row + 1, days_coord[1], pairs_coord[1], groups, all_merg)
                                for par in pair:
                                    all_timetable[fak][form][level][nap][info_table[0]][info_table[1]][par] = pair[par]
                            except:
                                print(info_table)
                                print(work_dir, exel)
                        else:
                            # pass
                            if "Медицинский" in work_dir:
                                print("spec")
                            special_group = []
                            while special_group == []:
                                try:
                                    group_row += 1
                                    # for col in range(pairs_coord[1]+1,sheet.max_column+1):
                                    start_coord = pairs_coord[1] + 1
                                    if not(isinstance(read_cell(sheet,group_row,days_coord[1],all_merg),datetime.datetime)) and\
                                        read_cell(sheet,group_row,days_coord[1],all_merg) is None:
                                        for index, name in enumerate(groups):
                                            val = read_cell(sheet, group_row, start_coord + index, all_merg)
                                            if any(j in str(val).lower() for j in ("день", "дата")):
                                                start_coord += 2
                                                val = read_cell(sheet, group_row, start_coord + index, all_merg)
                                            if any(j in str(val).lower() for j in ("пара", "время")):
                                                start_coord += 1
                                                val = read_cell(sheet, group_row, start_coord + index, all_merg)
                                            if re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\d+",
                                                          str(val)):  # число.буквы.число.буквы.число*число
                                                # val = re.sub(r"(\d)*[^0-9A-Я](\D*)[^0-9A-Я](\d*)[^0-9A-Я](\D*)[^0-9A-Я](\d*)[^0-9A-Я](\d*)",r"\1.\2.\3.\4.\5*\6",val)
                                                special_group.append(val)
                                            elif re.findall(r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\D+[^0-9A-Я]\d+",
                                                            str(val)):  # число.буквы.число.буквы.число
                                                special_group.append(val)
                                            elif re.findall(
                                                    r"\d+[^0-9A-Я]\D+[^0-9A-Я]\d+[^0-9A-Я]\d+",
                                                    str(val)):  # число.буквы.число.число
                                                special_group.append(val)
                                            elif str(val).isdigit():
                                                special_group.append(f"{name}*{val}")
                                            elif str(val) == name:
                                                special_group.append(val)
                                    else:
                                        group_row -=1
                                        break
                                except:
                                    print(work_dir,exel)
                                    exit()
                            if groups != [] and special_group ==[]:
                                special_group = groups
                            try:
                                pair = pars_line(sheet, group_row + 1, days_coord[1], pairs_coord[1], special_group,
                                                 all_merg)
                                for par in pair:
                                    all_timetable[fak][form][level][nap][info_table[0]][info_table[1]][par] = pair[par]
                            except:
                                print(info_table)
                                print(work_dir, exel)

    with open("all_timetable.json", "w") as file:
        json.dump(all_timetable, file, indent=3)
    print("Загрузка завешена")


if __name__ == '__main__':
    start = time.time()
    upload_rasp()
    print(time.time()-start)
